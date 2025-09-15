import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from botocore.exceptions import ClientError, EndpointConnectionError
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------
# CONFIG
# -------------------------
AWS_PROFILE = "aqcn"
INBOUND_FILE = "621924646669_security_groups_in_use_all_ports.xlsx"
OUTBOUND_FILE = "621924646669_security_groups_in_use_all_ports_outbound.xlsx"
MAX_THREADS = 5
# -------------------------

session = boto3.Session(profile_name=AWS_PROFILE)
ec2_client = session.client("ec2")

# ------------------- Excel Merge -------------------
def merge_cells_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        start_row = 2
        prev_val = ws[f"{col_letter}{start_row}"].value
        merge_start = start_row
        for row in range(start_row + 1, ws.max_row + 1):
            val = ws[f"{col_letter}{row}"].value
            if val != prev_val:
                if merge_start < row - 1 and prev_val is not None:
                    ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{row - 1}")
                    ws[f"{col_letter}{merge_start}"].alignment = Alignment(vertical="center", horizontal="center")
                merge_start = row
                prev_val = val
        if merge_start < ws.max_row and prev_val is not None:
            ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{ws.max_row}")
            ws[f"{col_letter}{merge_start}"].alignment = Alignment(vertical="center", horizontal="center")
    wb.save(file_path)

# ------------------- Collect all rules -------------------
def collect_rules(sg, direction="inbound"):
    results = []
    try:
        rules = sg["IpPermissions"] if direction == "inbound" else sg["IpPermissionsEgress"]
        for rule in rules:
            proto = rule.get("IpProtocol")
            fport = rule.get("FromPort")
            tport = rule.get("ToPort")
            for ip_range in rule.get("IpRanges", []):
                results.append((proto, fport, tport, ip_range["CidrIp"], "IPv4", ip_range.get("Description", "")))
            for ip_range in rule.get("Ipv6Ranges", []):
                results.append((proto, fport, tport, ip_range["CidrIpv6"], "IPv6", ip_range.get("Description", "")))
    except Exception as e:
        print(f"Error collecting {direction} rules: {e}")
    return results

# ------------------- Scan One Region -------------------
def scan_region(region):
    results = []
    try:
        ec2 = session.client("ec2", region_name=region)
        sgs = ec2.describe_security_groups()["SecurityGroups"]
        sg_map = {sg["GroupId"]: [] for sg in sgs}

        # ---------- ENI-first (captures most services) ----------
        try:
            paginator = ec2.get_paginator("describe_network_interfaces")
            for page in paginator.paginate():
                for eni in page["NetworkInterfaces"]:
                    desc = eni.get("Description", "")
                    name = next((t["Value"] for t in eni.get("TagSet", []) if t["Key"] == "Name"), desc)
                    for sg in eni.get("Groups", []):
                        sgid = sg["GroupId"]
                        if sgid in sg_map:
                            sg_map[sgid].append(("ENI", eni["NetworkInterfaceId"], name))
        except Exception as e:
            print(f"[ENI] {region} {e}")

        # ---------- EC2 ----------
        try:
            paginator = ec2.get_paginator("describe_instances")
            for page in paginator.paginate():
                for res in page["Reservations"]:
                    for inst in res["Instances"]:
                        name = next((t["Value"] for t in inst.get("Tags", []) if t["Key"] == "Name"), "")
                        for sg in inst.get("SecurityGroups", []):
                            sgid = sg["GroupId"]
                            if sgid in sg_map:
                                sg_map[sgid].append(("EC2", inst["InstanceId"], name))
        except Exception as e:
            print(f"[EC2] {region} {e}")

        # ---------- Classic ELB ----------
        try:
            elb = session.client("elb", region_name=region)
            paginator = elb.get_paginator("describe_load_balancers")
            for page in paginator.paginate():
                for lb in page["LoadBalancerDescriptions"]:
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].append(("Classic ELB", lb["LoadBalancerName"], lb.get("DNSName", "")))
        except Exception as e:
            print(f"[ELB] {region} {e}")

        # ---------- ALB / NLB ----------
        try:
            elbv2 = session.client("elbv2", region_name=region)
            paginator = elbv2.get_paginator("describe_load_balancers")
            for page in paginator.paginate():
                for lb in page["LoadBalancers"]:
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].append(("ALB/NLB", lb["LoadBalancerName"], lb.get("DNSName", "")))
        except Exception as e:
            print(f"[ALB/NLB] {region} {e}")

        # ---------- All other services (RDS, Redshift, ElastiCache, EFS, MQ, SageMaker, Lambda, MSK, Glue, App Runner) ----------
        # Use same logic as previous “ultimate” script
        # ENI-first scan already captures indirect SG associations

        # ---------- Collect Rules ----------
        for sg in sgs:
            sg_id = sg["GroupId"]
            sg_name = sg.get("GroupName", "")
            sg_desc = sg.get("Description", "")
            resources = sg_map.get(sg_id, [])
            if not resources:
                continue
            inbound_rules = collect_rules(sg, "inbound")
            outbound_rules = collect_rules(sg, "outbound")
            for res_type, res_id, res_name in resources:
                for proto, fport, tport, cidr, ip_ver, desc in inbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc, res_type, res_id, res_name,
                                    proto, fport, tport, cidr, ip_ver, desc, "INBOUND"])
                for proto, fport, tport, cidr, ip_ver, desc in outbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc, res_type, res_id, res_name,
                                    proto, fport, tport, cidr, ip_ver, desc, "OUTBOUND"])

    except EndpointConnectionError:
        print(f"Region {region} not available")
    except Exception as e:
        print(f"[{region}] Unexpected error: {e}")

    return results

# ------------------- Main -------------------
def main():
    inbound_data = []
    outbound_data = []

    regions = [r["RegionName"] for r in ec2_client.describe_regions()["Regions"]]

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        future_to_region = {executor.submit(scan_region, region): region for region in regions}
        for future in as_completed(future_to_region):
            region = future_to_region[future]
            try:
                results = future.result()
                for row in results:
                    if row[-1] == "INBOUND":
                        inbound_data.append(row[:-1])
                    else:
                        outbound_data.append(row[:-1])
                print(f"Completed region: {region}")
            except Exception as e:
                print(f"Error processing region {region}: {e}")

    columns = ["Region", "SG ID", "SG Name", "SG Description", "Resource Type", "Resource ID", "Resource Name",
               "Protocol", "FromPort", "ToPort", "CIDR", "IP Version", "Rule Description"]

    inbound_df = pd.DataFrame(inbound_data, columns=columns)
    outbound_df = pd.DataFrame(outbound_data, columns=columns)

    inbound_df.to_excel(INBOUND_FILE, index=False)
    outbound_df.to_excel(OUTBOUND_FILE, index=False)

    merge_cells_excel(INBOUND_FILE)
    merge_cells_excel(OUTBOUND_FILE)
    print(f"Exported results to {INBOUND_FILE} and {OUTBOUND_FILE}")

if __name__ == "__main__":
    main()
