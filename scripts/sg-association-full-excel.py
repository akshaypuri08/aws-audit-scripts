import boto3
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ---------------- Logging Setup ----------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

PROFILE_NAME = "aqcn"
OUTPUT_INBOUND = "sg_inbound.xlsx"
OUTPUT_OUTBOUND = "sg_outbound.xlsx"

# ---------------- AWS Session ----------------
try:
    boto3.setup_default_session(profile_name=PROFILE_NAME)
    session = boto3.Session(profile_name=PROFILE_NAME)
    logging.info(f"Using AWS profile: {PROFILE_NAME}")
except Exception as e:
    logging.error(f"Failed to create AWS session: {e}")
    raise

# ---------------- Get Regions ----------------
try:
    ec2_client = session.client("ec2")
    regions = [r["RegionName"] for r in ec2_client.describe_regions()["Regions"]]
    logging.info(f"Found {len(regions)} AWS regions: {regions}")
except Exception as e:
    logging.error(f"Failed to fetch regions: {e}")
    raise

csv_header = [
    "Region", "SecurityGroupId", "SecurityGroupName", "SecurityGroupDescription",
    "Protocol", "PortRange", "CIDR", "IP_Version",
    "RuleDescription",
    "Resource_Type", "ResourceId", "Resource_Name"
]

# Separate lists for inbound & outbound
inbound_results = []
outbound_results = []

def get_name_from_tags(tags):
    if not tags:
        return ""
    for tag in tags:
        if tag["Key"].lower() == "name":
            return tag["Value"]
    return ""

for region in regions:
    try:
        logging.info(f"Scanning region: {region}")
        ec2 = session.client("ec2", region_name=region)
        sgs = ec2.describe_security_groups()["SecurityGroups"]
        nis = ec2.describe_network_interfaces()["NetworkInterfaces"]

        sg_resources = {sg["GroupId"]: [] for sg in sgs}

        # Attach ENIs
        for ni in nis:
            for group in ni["Groups"]:
                sg_resources[group["GroupId"]].append(("NetworkInterface", ni["NetworkInterfaceId"], get_name_from_tags(ni.get("TagSet", []))))

        # Attach EC2
        reservations = ec2.describe_instances()["Reservations"]
        for res in reservations:
            for inst in res["Instances"]:
                for sg in inst["SecurityGroups"]:
                    sg_resources[sg["GroupId"]].append(("EC2", inst["InstanceId"], get_name_from_tags(inst.get("Tags", []))))

        # Attach Load Balancers
        elb = session.client("elb", region_name=region)
        elbv2 = session.client("elbv2", region_name=region)
        for lb in elb.describe_load_balancers().get("LoadBalancerDescriptions", []):
            for sg in lb.get("SecurityGroups", []):
                sg_resources[sg].append(("ELB", lb["LoadBalancerName"], lb["LoadBalancerName"]))
        for lb in elbv2.describe_load_balancers().get("LoadBalancers", []):
            for sg in lb.get("SecurityGroups", []):
                sg_resources[sg].append(("ALB/NLB", lb["LoadBalancerArn"], lb["LoadBalancerName"]))

        # Process SGs
        for sg in sgs:
            sg_id = sg["GroupId"]
            sg_name = sg.get("GroupName", "")
            sg_desc = sg.get("Description", "")

            if not sg_resources.get(sg_id):
                continue  # Only attached SGs

            # ----------- INBOUND RULES -----------
            for perm in sg.get("IpPermissions", []):
                protocol = perm.get("IpProtocol", "-1")
                from_port = perm.get("FromPort", "All") if "FromPort" in perm else "All"
                to_port = perm.get("ToPort", "All") if "ToPort" in perm else "All"
                port_range = f"{from_port}-{to_port}" if from_port != to_port else str(from_port)

                for cidr in perm.get("IpRanges", []):
                    rule_desc = cidr.get("Description", "")
                    for res_type, res_id, res_name in sg_resources[sg_id]:
                        inbound_results.append([
                            region, sg_id, sg_name, sg_desc, protocol, port_range,
                            cidr.get("CidrIp", ""), "IPv4", rule_desc,
                            res_type, res_id, res_name
                        ])
                for cidr in perm.get("Ipv6Ranges", []):
                    rule_desc = cidr.get("Description", "")
                    for res_type, res_id, res_name in sg_resources[sg_id]:
                        inbound_results.append([
                            region, sg_id, sg_name, sg_desc, protocol, port_range,
                            cidr.get("CidrIpv6", ""), "IPv6", rule_desc,
                            res_type, res_id, res_name
                        ])

            # ----------- OUTBOUND RULES -----------
            for perm in sg.get("IpPermissionsEgress", []):
                protocol = perm.get("IpProtocol", "-1")
                from_port = perm.get("FromPort", "All") if "FromPort" in perm else "All"
                to_port = perm.get("ToPort", "All") if "ToPort" in perm else "All"
                port_range = f"{from_port}-{to_port}" if from_port != to_port else str(from_port)

                for cidr in perm.get("IpRanges", []):
                    rule_desc = cidr.get("Description", "")
                    for res_type, res_id, res_name in sg_resources[sg_id]:
                        outbound_results.append([
                            region, sg_id, sg_name, sg_desc, protocol, port_range,
                            cidr.get("CidrIp", ""), "IPv4", rule_desc,
                            res_type, res_id, res_name
                        ])
                for cidr in perm.get("Ipv6Ranges", []):
                    rule_desc = cidr.get("Description", "")
                    for res_type, res_id, res_name in sg_resources[sg_id]:
                        outbound_results.append([
                            region, sg_id, sg_name, sg_desc, protocol, port_range,
                            cidr.get("CidrIpv6", ""), "IPv6", rule_desc,
                            res_type, res_id, res_name
                        ])

        logging.info(f"Completed scanning region: {region}")

    except Exception as e:
        logging.error(f"Error processing region {region}: {e}")

# ---------------- Helper to Save Excel with Merge ----------------
def save_excel_with_merge(data, output_file):
    df = pd.DataFrame(data, columns=csv_header)
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    merge_columns = [
        "Region", "SecurityGroupId", "SecurityGroupName", "SecurityGroupDescription",
        "Resource_Type", "ResourceId", "Resource_Name"
    ]
    merge_col_indexes = [csv_header.index(col) + 1 for col in merge_columns]

    for col_idx in merge_col_indexes:
        col_letter = get_column_letter(col_idx)
        start_row = 2
        current_value = ws[f"{col_letter}{start_row}"].value
        for row in range(3, ws.max_row + 2):
            cell_value = ws[f"{col_letter}{row}"].value if row <= ws.max_row else None
            if cell_value != current_value:
                end_row = row - 1
                if end_row > start_row:
                    ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                start_row = row
                current_value = cell_value

    wb.save(output_file)
    logging.info(f"Excel file saved: {output_file}")

# Save inbound & outbound separately
if inbound_results:
    save_excel_with_merge(inbound_results, OUTPUT_INBOUND)
else:
    logging.warning("No inbound rules found")

if outbound_results:
    save_excel_with_merge(outbound_results, OUTPUT_OUTBOUND)
else:
    logging.warning("No outbound rules found")
