import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from botocore.exceptions import ClientError, EndpointConnectionError
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------
# CONFIG
# -------------------------
AWS_PROFILE = "aqcn"   # Change this
TARGET_PORTS = {20, 21, 22, 3389, 5432}
INBOUND_FILE = "621924646669_security_groups_in_use_selected_ports.xlsx"
OUTBOUND_FILE = "621924646669_security_groups_in_use_selected_ports_outbound.xlsx"
MAX_THREADS = 5  # Number of regions scanned in parallel
# -------------------------

session = boto3.Session(profile_name=AWS_PROFILE)
ec2_client = session.client("ec2")

# ------------------- Excel Merge -------------------
def merge_cells_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        start_row = 2
        prev_val = ws[f"{col_letter}{start_row}"].value
        merge_start = start_row
        for row in range(start_row + 1, ws.max_row + 1):
            val = ws[f"{col_letter}{row}"].value
            if val != prev_val:
                if merge_start < row - 1 and prev_val is not None:
                    ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{row - 1}")
                    ws[f"{col_letter}{merge_start}"].alignment = Alignment(vertical="center", horizontal="center")
                merge_start = row
                prev_val = val
        if merge_start < ws.max_row and prev_val is not None:
            ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{ws.max_row}")
            ws[f"{col_letter}{merge_start}"].alignment = Alignment(vertical="center", horizontal="center")
    wb.save(file_path)

# ------------------- Rule Filtering -------------------
def rule_matches_ports(rule):
    from_port = rule.get("FromPort")
    to_port = rule.get("ToPort")
    if from_port is None or to_port is None:
        return False
    return any(from_port <= port <= to_port for port in TARGET_PORTS)

def collect_rules(sg, direction="inbound"):
    results = []
    try:
        rules = sg["IpPermissions"] if direction == "inbound" else sg["IpPermissionsEgress"]
        for rule in rules:
            if not rule_matches_ports(rule):
                continue
            proto = rule.get("IpProtocol")
            fport = rule.get("FromPort")
            tport = rule.get("ToPort")
            for ip_range in rule.get("IpRanges", []):
                results.append((proto, fport, tport, ip_range["CidrIp"], "IPv4", ip_range.get("Description", "")))
            for ip_range in rule.get("Ipv6Ranges", []):
                results.append((proto, fport, tport, ip_range["CidrIpv6"], "IPv6", ip_range.get("Description", "")))
    except Exception as e:
        print(f"Error collecting {direction} rules: {e}")
    return results

# ------------------- Scan One Region -------------------
def scan_region(region):
    results = []
    try:
        ec2 = session.client("ec2", region_name=region)
        sgs = ec2.describe_security_groups()["SecurityGroups"]
        sg_map = {sg["GroupId"]: [] for sg in sgs}

        # ---------- ENI-first (captures most services) ----------
        try:
            paginator = ec2.get_paginator("describe_network_interfaces")
            for page in paginator.paginate():
                for eni in page["NetworkInterfaces"]:
                    desc = eni.get("Description", "")
                    name = next((t["Value"] for t in eni.get("TagSet", []) if t["Key"] == "Name"), desc)
                    for sg in eni.get("Groups", []):
                        sgid = sg["GroupId"]
                        if sgid in sg_map:
                            sg_map[sgid].append(("ENI", eni["NetworkInterfaceId"], name))
        except Exception as e:
            print(f"[ENI] {region} {e}")

        # ---------- EC2 Instances ----------
        try:
            paginator = ec2.get_paginator("describe_instances")
            for page in paginator.paginate():
                for res in page["Reservations"]:
                    for inst in res["Instances"]:
                        name = next((t["Value"] for t in inst.get("Tags", []) if t["Key"] == "Name"), "")
                        for sg in inst.get("SecurityGroups", []):
                            sgid = sg["GroupId"]
                            if sgid in sg_map:
                                sg_map[sgid].append(("EC2", inst["InstanceId"], name))
        except Exception as e:
            print(f"[EC2] {region} {e}")

        # ---------- Classic ELB ----------
        try:
            elb = session.client("elb", region_name=region)
            paginator = elb.get_paginator("describe_load_balancers")
            for page in paginator.paginate():
                for lb in page["LoadBalancerDescriptions"]:
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].append(("Classic ELB", lb["LoadBalancerName"], lb.get("DNSName", "")))
        except Exception as e:
            print(f"[ELB] {region} {e}")

        # ---------- ALB / NLB ----------
        try:
            elbv2 = session.client("elbv2", region_name=region)
            paginator = elbv2.get_paginator("describe_load_balancers")
            for page in paginator.paginate():
                for lb in page["LoadBalancers"]:
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].append(("ALB/NLB", lb["LoadBalancerName"], lb.get("DNSName", "")))
        except Exception as e:
            print(f"[ALB/NLB] {region} {e}")

        # ---------- RDS ----------
        try:
            rds = session.client("rds", region_name=region)
            for db in rds.describe_db_instances()["DBInstances"]:
                for vpcsg in db.get("VpcSecurityGroups", []):
                    sgid = vpcsg["VpcSecurityGroupId"]
                    if sgid in sg_map:
                        name = next((t["Value"] for t in db.get("TagList", []) if t["Key"] == "Name"), "")
                        sg_map[sgid].append(("RDS Instance", db["DBInstanceIdentifier"], name))
            for cluster in rds.describe_db_clusters()["DBClusters"]:
                for vpcsg in cluster.get("VpcSecurityGroups", []):
                    sgid = vpcsg["VpcSecurityGroupId"]
                    if sgid in sg_map:
                        sg_map[sgid].append(("RDS Cluster", cluster["DBClusterIdentifier"], ""))
        except Exception as e:
            print(f"[RDS] {region} {e}")

        # ---------- ElastiCache ----------
        try:
            elasticache = session.client("elasticache", region_name=region)
            for cl in elasticache.describe_cache_clusters(ShowCacheNodeInfo=True)["CacheClusters"]:
                for vpcsg in cl.get("SecurityGroups", []):
                    sgid = vpcsg["SecurityGroupId"]
                    if sgid in sg_map:
                        sg_map[sgid].append(("ElastiCache", cl["CacheClusterId"], ""))
        except Exception as e:
            print(f"[ElastiCache] {region} {e}")

        # ---------- Redshift ----------
        try:
            redshift = session.client("redshift", region_name=region)
            for cluster in redshift.describe_clusters()["Clusters"]:
                for vpcsg in cluster.get("VpcSecurityGroups", []):
                    sgid = vpcsg["VpcSecurityGroupId"]
                    if sgid in sg_map:
                        sg_map[sgid].append(("Redshift", cluster["ClusterIdentifier"], ""))
        except Exception as e:
            print(f"[Redshift] {region} {e}")

        # ---------- EFS ----------
        try:
            efs = session.client("efs", region_name=region)
            for fs in efs.describe_file_systems()["FileSystems"]:
                for m in efs.describe_mount_targets(FileSystemId=fs["FileSystemId"])["MountTargets"]:
                    sg_list = efs.describe_mount_target_security_groups(MountTargetId=m["MountTargetId"])["SecurityGroups"]
                    for sgid in sg_list:
                        if sgid in sg_map:
                            sg_map[sgid].append(("EFS", fs["FileSystemId"], ""))
        except Exception as e:
            print(f"[EFS] {region} {e}")

        # ---------- MQ ----------
        try:
            mq = session.client("mq", region_name=region)
            for broker in mq.list_brokers()["BrokerSummaries"]:
                details = mq.describe_broker(BrokerId=broker["BrokerId"])
                for sgid in details.get("SecurityGroups", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("MQ", broker["BrokerName"], ""))
        except Exception as e:
            print(f"[MQ] {region} {e}")

        # ---------- SageMaker ----------
        try:
            sm = session.client("sagemaker", region_name=region)
            for nb in sm.list_notebook_instances()["NotebookInstances"]:
                for sgid in sm.describe_notebook_instance(NotebookInstanceName=nb["NotebookInstanceName"]).get("SecurityGroups", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("SageMaker Notebook", nb["NotebookInstanceName"], ""))
            for ep in sm.list_endpoints()["Endpoints"]:
                for sgid in sm.describe_endpoint(EndpointName=ep["EndpointName"]).get("VpcConfig", {}).get("SecurityGroupIds", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("SageMaker Endpoint", ep["EndpointName"], ""))
        except Exception as e:
            print(f"[SageMaker] {region} {e}")

        # ---------- Lambda ----------
        try:
            lambda_client = session.client("lambda", region_name=region)
            for fn in lambda_client.list_functions()["Functions"]:
                cfg = lambda_client.get_function_configuration(FunctionName=fn["FunctionName"])
                for sgid in cfg.get("VpcConfig", {}).get("SecurityGroupIds", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("Lambda", fn["FunctionName"], ""))
        except Exception as e:
            print(f"[Lambda] {region} {e}")

        # ---------- MSK ----------
        try:
            msk = session.client("kafka", region_name=region)
            for cluster in msk.list_clusters()["ClusterInfoList"]:
                details = msk.describe_cluster(ClusterArn=cluster["ClusterArn"])["ClusterInfo"]
                for sgid in details.get("BrokerNodeGroupInfo", {}).get("SecurityGroups", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("MSK", cluster["ClusterName"], ""))
        except Exception as e:
            print(f"[MSK] {region} {e}")

        # ---------- Glue ----------
        try:
            glue = session.client("glue", region_name=region)
            for dev in glue.get_dev_endpoints()["DevEndpoints"]:
                for sgid in dev.get("SecurityGroupIds", []):
                    if sgid in sg_map:
                        sg_map[sgid].append(("Glue Dev Endpoint", dev["EndpointName"], ""))
        except Exception as e:
            print(f"[Glue] {region} {e}")

        # ---------- App Runner ----------
        try:
            apprunner = session.client("apprunner", region_name=region)
            for svc in apprunner.list_services()["ServiceSummaryList"]:
                details = apprunner.describe_service(ServiceArn=svc["ServiceArn"])["Service"]
                vpc_conn = details.get("NetworkConfiguration", {}).get("VpcConnectorArn", [])
                if isinstance(vpc_conn, str):
                    sgid = vpc_conn  # For simplicity
                    if sgid in sg_map:
                        sg_map[sgid].append(("App Runner", svc["ServiceName"], ""))
        except Exception as e:
            print(f"[AppRunner] {region} {e}")

        # ---------- Remaining services (Transit Gateway, Neptune, DocumentDB, OpenSearch, DMS, WorkSpaces, AppStream, DataSync) ----------
        # Already captured via ENI scan (ENI-first strategy)

        # ---------- Collect Rules ----------
        for sg in sgs:
            sg_id = sg["GroupId"]
            sg_name = sg.get("GroupName", "")
            sg_desc = sg.get("Description", "")
            resources = sg_map.get(sg_id, [])
            if not resources:
                continue
            inbound_rules = collect_rules(sg, "inbound")
            outbound_rules = collect_rules(sg, "outbound")
            for res_type, res_id, res_name in resources:
                for proto, fport, tport, cidr, ip_ver, desc in inbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc, res_type, res_id, res_name,
                                    proto, fport, tport, cidr, ip_ver, desc, "INBOUND"])
                for proto, fport, tport, cidr, ip_ver, desc in outbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc, res_type, res_id, res_name,
                                    proto, fport, tport, cidr, ip_ver, desc, "OUTBOUND"])

    except EndpointConnectionError:
        print(f"Region {region} not available")
    except Exception as e:
        print(f"[{region}] Unexpected error: {e}")

    return results

# ------------------- Main -------------------
def main():
    inbound_data = []
    outbound_data = []

    regions = [r["RegionName"] for r in ec2_client.describe_regions()["Regions"]]

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        future_to_region = {executor.submit(scan_region, region): region for region in regions}
        for future in as_completed(future_to_region):
            region = future_to_region[future]
            try:
                results = future.result()
                for row in results:
                    if row[-1] == "INBOUND":
                        inbound_data.append(row[:-1])
                    else:
                        outbound_data.append(row[:-1])
                print(f"Completed region: {region}")
            except Exception as e:
                print(f"Error processing region {region}: {e}")

    columns = ["Region", "SG ID", "SG Name", "SG Description", "Resource Type", "Resource ID", "Resource Name",
               "Protocol", "FromPort", "ToPort", "CIDR", "IP Version", "Rule Description"]

    inbound_df = pd.DataFrame(inbound_data, columns=columns)
    outbound_df = pd.DataFrame(outbound_data, columns=columns)

    inbound_df.to_excel(INBOUND_FILE, index=False)
    outbound_df.to_excel(OUTBOUND_FILE, index=False)

    merge_cells_excel(INBOUND_FILE)
    merge_cells_excel(OUTBOUND_FILE)
    print(f"Exported results to {INBOUND_FILE} and {OUTBOUND_FILE}")

if __name__ == "__main__":
    main()
