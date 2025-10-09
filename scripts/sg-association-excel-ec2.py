import boto3
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from botocore.exceptions import EndpointConnectionError
from botocore.config import Config
from concurrent.futures import ThreadPoolExecutor, as_completed

AWS_PROFILE = "aqcn"
INBOUND_FILE = "ec2_security_groups_inbound.xlsx"
OUTBOUND_FILE = "ec2_security_groups_outbound.xlsx"
MAX_THREADS = 5

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO
)

session = boto3.Session(profile_name=AWS_PROFILE)
config = Config(retries={'max_attempts': 5, 'mode': 'adaptive'}, connect_timeout=5, read_timeout=10)
ec2_client = session.client("ec2", config=config)

# ---------------- Utilities ----------------
def normalize_protocol(proto):
    if proto is None:
        return ""
    p = str(proto)
    return "ALL" if p == "-1" else p

def format_port_range(proto, fport, tport):
    p = normalize_protocol(proto)
    if p == "ALL":
        return "ALL"
    lowerp = p.lower()
    if lowerp.startswith("icmp"):
        if fport is None and tport is None:
            return ""
        if fport is None:
            return str(tport)
        if tport is None:
            return str(fport)
        return f"{fport}/{tport}"
    if fport is None and tport is None:
        return ""
    return str(fport) if fport == tport else f"{fport}-{tport}"

def collect_rules(sg, direction="inbound"):
    results = []
    try:
        rules = sg.get("IpPermissions") if direction == "inbound" else sg.get("IpPermissionsEgress", [])
        for rule in rules:
            proto_raw = rule.get("IpProtocol")
            proto = normalize_protocol(proto_raw)
            port_str = format_port_range(proto_raw, rule.get("FromPort"), rule.get("ToPort"))
            for ip_range in rule.get("IpRanges", []):
                results.append((proto, port_str, ip_range.get("CidrIp"), "IPv4", ip_range.get("Description", "")))
            for ip_range in rule.get("Ipv6Ranges", []):
                results.append((proto, port_str, ip_range.get("CidrIpv6"), "IPv6", ip_range.get("Description", "")))
            for pl in rule.get("PrefixListIds", []):
                results.append((proto, port_str, pl.get("PrefixListId"), "PrefixList", pl.get("Description", "")))
            for pair in rule.get("UserIdGroupPairs", []):
                display = f"sg:{pair.get('GroupId')}" if pair.get("GroupId") else pair.get("UserId", "")
                results.append((proto, port_str, display, "SecurityGroup", pair.get("Description", "")))
    except Exception as e:
        logging.error(f"Error collecting {direction} rules: {e}")
    return results

def merge_selected_columns(file_path, merge_columns):
    wb = load_workbook(file_path)
    ws = wb.active
    col_index_map = {col.value: idx + 1 for idx, col in enumerate(ws[1]) if col.value in merge_columns}

    for col_name, col_idx in col_index_map.items():
        start_row = 2
        prev_val = ws.cell(row=start_row, column=col_idx).value
        merge_start = start_row

        for row in range(start_row + 1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val != prev_val:
                if merge_start < row - 1 and prev_val is not None:
                    ws.merge_cells(start_row=merge_start, start_column=col_idx,
                                   end_row=row - 1, end_column=col_idx)
                    ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical="center", horizontal="center")
                merge_start = row
                prev_val = val

        if merge_start < ws.max_row and prev_val is not None:
            ws.merge_cells(start_row=merge_start, start_column=col_idx,
                           end_row=ws.max_row, end_column=col_idx)
            ws.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical="center", horizontal="center")

    wb.save(file_path)

# ---------------- Main Region Scanner ----------------
def scan_region(region):
    results = []
    try:
        ec2 = session.client("ec2", region_name=region, config=config)

        # -------- Build Subnet Map --------
        subnet_map = {}
        try:
            subnets = []
            for page in ec2.get_paginator("describe_subnets").paginate():
                subnets.extend(page["Subnets"])

            # Get route tables for correct public/private classification
            vpc_ids = list({s["VpcId"] for s in subnets})
            route_tables = []
            for vpc_id in vpc_ids:
                for page in ec2.get_paginator("describe_route_tables").paginate(Filters=[{"Name": "vpc-id", "Values": [vpc_id]}]):
                    route_tables.extend(page["RouteTables"])

            subnet_to_rtb, main_rtb_per_vpc = {}, {}
            for rtb in route_tables:
                for assoc in rtb.get("Associations", []):
                    if assoc.get("Main"):
                        main_rtb_per_vpc[rtb["VpcId"]] = rtb
                    elif assoc.get("SubnetId"):
                        subnet_to_rtb[assoc["SubnetId"]] = rtb

            for subnet in subnets:
                subnet_id = subnet["SubnetId"]
                subnet_name = next((t["Value"] for t in subnet.get("Tags", []) if t["Key"] == "Name"), "")
                rtb = subnet_to_rtb.get(subnet_id) or main_rtb_per_vpc.get(subnet["VpcId"])
                subnet_type = "Private"
                if rtb:
                    for route in rtb.get("Routes", []):
                        if (route.get("GatewayId") or "").startswith("igw-"):
                            subnet_type = "Public"
                            break
                subnet_map[subnet_id] = (subnet_name, subnet_type)
        except Exception as e:
            logging.warning(f"[Subnet] {region} {e}")

        # -------- Get Security Groups --------
        sgs = ec2.describe_security_groups().get("SecurityGroups", [])
        sg_map = {sg["GroupId"]: [] for sg in sgs}

        # -------- Get EC2 Instances and Map SGs --------
        try:
            paginator = ec2.get_paginator("describe_instances")
            for page in paginator.paginate():
                for res in page["Reservations"]:
                    for inst in res["Instances"]:
                        instance_id = inst["InstanceId"]
                        instance_name = next((t["Value"] for t in inst.get("Tags", []) if t["Key"] == "Name"), "")
                        subnet_id = inst.get("SubnetId", "")
                        subnet_name, subnet_type = subnet_map.get(subnet_id, ("", ""))

                        for sg in inst.get("SecurityGroups", []):
                            sgid = sg["GroupId"]
                            if sgid in sg_map:
                                sg_map[sgid].append((instance_id, instance_name, subnet_id, subnet_name, subnet_type))
        except Exception as e:
            logging.warning(f"[EC2] {region} {e}")

        # -------- Collect Rules --------
        for sg in sgs:
            sg_id = sg["GroupId"]
            sg_name = sg.get("GroupName", "")
            sg_desc = sg.get("Description", "")
            resources = sg_map.get(sg_id, [])
            if not resources:
                continue
            inbound_rules = collect_rules(sg, "inbound")
            outbound_rules = collect_rules(sg, "outbound")
            for res in resources:
                instance_id, instance_name, subnet_id, subnet_name, subnet_type = res
                for proto, port_str, cidr, ip_ver, desc in inbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc,
                                    instance_id, instance_name,
                                    subnet_name, subnet_id, subnet_type,
                                    proto, port_str, cidr, ip_ver, desc, "INBOUND"])
                for proto, port_str, cidr, ip_ver, desc in outbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc,
                                    instance_id, instance_name,
                                    subnet_name, subnet_id, subnet_type,
                                    proto, port_str, cidr, ip_ver, desc, "OUTBOUND"])
    except EndpointConnectionError:
        logging.warning(f"Region {region} not available")
    except Exception as e:
        logging.error(f"[{region}] Unexpected error: {e}")

    return results

# ---------------- Main ----------------
def main():
    inbound_data, outbound_data = [], []
    regions = [r["RegionName"] for r in ec2_client.describe_regions()["Regions"]]

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        future_to_region = {executor.submit(scan_region, region): region for region in regions}
        for future in as_completed(future_to_region):
            region = future_to_region[future]
            try:
                results = future.result()
                for row in results:
                    if row[-1] == "INBOUND":
                        inbound_data.append(row[:-1])
                    else:
                        outbound_data.append(row[:-1])
                logging.info(f"Completed region: {region}")
            except Exception as e:
                logging.error(f"Error processing region {region}: {e}")

    columns = ["Region", "SG ID", "SG Name", "SG Description",
               "Instance ID", "Instance Name",
               "Subnet Name", "Subnet ID", "Subnet Type",
               "Protocol", "Port(s)", "CIDR/Reference", "IP Version/Type", "Rule Description"]

    inbound_df = pd.DataFrame(inbound_data, columns=columns)
    outbound_df = pd.DataFrame(outbound_data, columns=columns)

    inbound_df.to_excel(INBOUND_FILE, index=False)
    outbound_df.to_excel(OUTBOUND_FILE, index=False)

    merge_cols = ["Region", "SG ID", "SG Name", "SG Description",
                  "Instance ID", "Instance Name", "Subnet Name", "Subnet ID"]
    merge_selected_columns(INBOUND_FILE, merge_cols)
    merge_selected_columns(OUTBOUND_FILE, merge_cols)

    logging.info(f"Exported results to {INBOUND_FILE} and {OUTBOUND_FILE}")

if __name__ == "__main__":
    main()
