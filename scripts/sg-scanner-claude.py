import boto3
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from botocore.exceptions import EndpointConnectionError
from botocore.config import Config
from concurrent.futures import ThreadPoolExecutor, as_completed

# -------------------------
# CONFIG
# -------------------------
AWS_PROFILE = "linko"
MAX_THREADS = 5
# -------------------------

# Configure logging
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO
)


# ------------------- Helpers -------------------

def get_account_id(session):
    try:
        return session.client("sts").get_caller_identity()["Account"]
    except Exception as e:
        logging.warning(f"Could not retrieve account ID: {e}")
        return "unknown"


def normalize_protocol(proto):
    if proto is None:
        return ""
    p = str(proto)
    if p == "-1":
        return "ALL"
    return p


def format_port_range(proto, fport, tport):
    p = normalize_protocol(proto)
    if p == "ALL":
        return "ALL"

    lowerp = p.lower()
    if lowerp.startswith("icmp"):
        if fport is None and tport is None:
            return ""
        if fport is None:
            return str(tport)
        if tport is None:
            return str(fport)
        return f"{fport}/{tport}"

    if fport is None and tport is None:
        return ""
    if fport == tport:
        return str(fport)
    return f"{fport}-{tport}"


def collect_rules(sg, direction="inbound"):
    results = []
    try:
        rules = sg.get("IpPermissions") if direction == "inbound" else sg.get("IpPermissionsEgress", [])
        for rule in rules:
            proto_raw = rule.get("IpProtocol")
            proto = normalize_protocol(proto_raw)
            port_str = format_port_range(proto_raw, rule.get("FromPort"), rule.get("ToPort"))

            for ip_range in rule.get("IpRanges", []):
                results.append((proto, port_str, ip_range.get("CidrIp"), "IPv4", ip_range.get("Description", "")))
            for ip_range in rule.get("Ipv6Ranges", []):
                results.append((proto, port_str, ip_range.get("CidrIpv6"), "IPv6", ip_range.get("Description", "")))
            for pl in rule.get("PrefixListIds", []):
                results.append((proto, port_str, pl.get("PrefixListId"), "PrefixList", pl.get("Description", "")))
            for pair in rule.get("UserIdGroupPairs", []):
                display = f"sg:{pair.get('GroupId')}" if pair.get("GroupId") else pair.get("UserId", "")
                results.append((proto, port_str, display, "SecurityGroup", pair.get("Description", "")))

    except Exception as e:
        logging.error(f"Error collecting {direction} rules: {e}")
    return results


# ------------------- Excel Formatting -------------------

def apply_excel_formatting(file_path, merge_columns):
    wb = load_workbook(file_path)
    ws = wb.active

    # Header styling
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto column widths
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Merge selected columns
    col_index_map = {col.value: idx + 1 for idx, col in enumerate(ws[1]) if col.value in merge_columns}

    for col_name, col_idx in col_index_map.items():
        start_row = 2
        prev_val = ws.cell(row=start_row, column=col_idx).value
        merge_start = start_row

        for row in range(start_row + 1, ws.max_row + 2):  # +2 to flush final group
            val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else object()  # sentinel
            if val != prev_val:
                if merge_start < row - 1 and prev_val is not None:
                    ws.merge_cells(
                        start_row=merge_start, start_column=col_idx,
                        end_row=row - 1, end_column=col_idx
                    )
                    ws.cell(row=merge_start, column=col_idx).alignment = Alignment(
                        vertical="center", horizontal="center", wrap_text=True
                    )
                merge_start = row
                prev_val = val

    wb.save(file_path)


# ------------------- Scan One Region -------------------

def scan_region(region, aws_profile):
    results = []
    orphaned_sgs = []

    try:
        session = boto3.Session(profile_name=aws_profile)  # Thread-safe: new session per thread
        config = Config(retries={'max_attempts': 5, 'mode': 'adaptive'}, connect_timeout=5, read_timeout=10)
        ec2 = session.client("ec2", region_name=region, config=config)

        # ---- Subnet Metadata ----
        subnet_map = {}
        try:
            subnets = []
            for page in ec2.get_paginator("describe_subnets").paginate():
                subnets.extend(page["Subnets"])

            vpc_ids = list({s["VpcId"] for s in subnets}) if subnets else []
            route_tables = []
            for vpc_id in vpc_ids:
                for page in ec2.get_paginator("describe_route_tables").paginate(
                    Filters=[{"Name": "vpc-id", "Values": [vpc_id]}]
                ):
                    route_tables.extend(page["RouteTables"])

            subnet_to_rtb, main_rtb_per_vpc = {}, {}
            for rtb in route_tables:
                for assoc in rtb.get("Associations", []):
                    if assoc.get("Main"):
                        main_rtb_per_vpc[rtb["VpcId"]] = rtb
                    elif assoc.get("SubnetId"):
                        subnet_to_rtb[assoc["SubnetId"]] = rtb

            for subnet in subnets:
                subnet_id = subnet["SubnetId"]
                subnet_name = next((t["Value"] for t in subnet.get("Tags", []) if t["Key"] == "Name"), "")
                rtb = subnet_to_rtb.get(subnet_id) or main_rtb_per_vpc.get(subnet["VpcId"])
                subnet_type = "Private"
                if rtb:
                    for route in rtb.get("Routes", []):
                        if (route.get("GatewayId") or "").startswith("igw-"):
                            subnet_type = "Public"
                            break
                subnet_map[subnet_id] = (subnet_name, subnet_type)
        except Exception as e:
            logging.warning(f"[Subnet] {region}: {e}")

        # ---- Security Groups (paginated) ----
        sgs = []
        try:
            for page in ec2.get_paginator("describe_security_groups").paginate():
                sgs.extend(page["SecurityGroups"])
        except Exception as e:
            logging.warning(f"[SG] {region}: {e}")
            return results, orphaned_sgs

        sg_map = {sg["GroupId"]: set() for sg in sgs}  # set of tuples to deduplicate

        # ---- ENI ----
        try:
            for page in ec2.get_paginator("describe_network_interfaces").paginate():
                for eni in page["NetworkInterfaces"]:
                    subnet_id = eni.get("SubnetId", "")
                    subnet_name, subnet_type = subnet_map.get(subnet_id, ("", ""))
                    desc = eni.get("Description", "")
                    name = next((t["Value"] for t in eni.get("TagSet", []) if t["Key"] == "Name"), desc)
                    for sg in eni.get("Groups", []):
                        sgid = sg["GroupId"]
                        if sgid in sg_map:
                            sg_map[sgid].add(("ENI", eni["NetworkInterfaceId"], name, subnet_id, subnet_name, subnet_type))
        except Exception as e:
            logging.warning(f"[ENI] {region}: {e}")

        # ---- EC2 ----
        try:
            for page in ec2.get_paginator("describe_instances").paginate():
                for res in page["Reservations"]:
                    for inst in res["Instances"]:
                        subnet_id = inst.get("SubnetId", "")
                        subnet_name, subnet_type = subnet_map.get(subnet_id, ("", ""))
                        name = next((t["Value"] for t in inst.get("Tags", []) if t["Key"] == "Name"), "")
                        for sg in inst.get("SecurityGroups", []):
                            sgid = sg["GroupId"]
                            if sgid in sg_map:
                                sg_map[sgid].add(("EC2", inst["InstanceId"], name, subnet_id, subnet_name, subnet_type))
        except Exception as e:
            logging.warning(f"[EC2] {region}: {e}")

        # ---- Classic ELB ----
        try:
            elb = session.client("elb", region_name=region, config=config)
            for page in elb.get_paginator("describe_load_balancers").paginate():
                for lb in page.get("LoadBalancerDescriptions", []):
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].add(("Classic ELB", lb["LoadBalancerName"], lb.get("DNSName", ""), "", "", ""))
        except Exception as e:
            logging.warning(f"[ELB] {region}: {e}")

        # ---- ALB/NLB ----
        try:
            elbv2 = session.client("elbv2", region_name=region, config=config)
            for page in elbv2.get_paginator("describe_load_balancers").paginate():
                for lb in page.get("LoadBalancers", []):
                    for sgid in lb.get("SecurityGroups", []):
                        if sgid in sg_map:
                            sg_map[sgid].add(("ALB/NLB", lb["LoadBalancerName"], lb.get("DNSName", ""), "", "", ""))
        except Exception as e:
            logging.warning(f"[ALB/NLB] {region}: {e}")

        # ---- RDS ----
        try:
            rds = session.client("rds", region_name=region, config=config)
            for page in rds.get_paginator("describe_db_instances").paginate():
                for db in page["DBInstances"]:
                    subnet_id = ""
                    subnet_name, subnet_type = "", ""
                    for vsg in db.get("VpcSecurityGroups", []):
                        sgid = vsg.get("VpcSecurityGroupId")
                        if sgid in sg_map:
                            sg_map[sgid].add(("RDS", db["DBInstanceIdentifier"], db.get("DBName", ""), subnet_id, subnet_name, subnet_type))
        except Exception as e:
            logging.warning(f"[RDS] {region}: {e}")

        # ---- Lambda ----
        try:
            lmb = session.client("lambda", region_name=region, config=config)
            paginator = lmb.get_paginator("list_functions")
            for page in paginator.paginate():
                for fn in page["Functions"]:
                    vpc_config = fn.get("VpcConfig", {})
                    for sgid in vpc_config.get("SecurityGroupIds", []):
                        if sgid in sg_map:
                            subnet_id = (vpc_config.get("SubnetIds") or [""])[0]
                            subnet_name, subnet_type = subnet_map.get(subnet_id, ("", ""))
                            sg_map[sgid].add(("Lambda", fn["FunctionName"], fn.get("Description", ""), subnet_id, subnet_name, subnet_type))
        except Exception as e:
            logging.warning(f"[Lambda] {region}: {e}")

        # ---- ElastiCache ----
        try:
            ec_client = session.client("elasticache", region_name=region, config=config)
            paginator = ec_client.get_paginator("describe_cache_clusters")
            for page in paginator.paginate(ShowCacheNodeInfo=False):
                for cluster in page["CacheClusters"]:
                    for sg in cluster.get("SecurityGroups", []):
                        sgid = sg.get("SecurityGroupId")
                        if sgid in sg_map:
                            sg_map[sgid].add(("ElastiCache", cluster["CacheClusterId"], cluster.get("Engine", ""), "", "", ""))
        except Exception as e:
            logging.warning(f"[ElastiCache] {region}: {e}")

        # ---- ECS ----
        try:
            ecs = session.client("ecs", region_name=region, config=config)
            clusters = ecs.list_clusters().get("clusterArns", [])
            for cluster_arn in clusters:
                task_arns = []
                paginator = ecs.get_paginator("list_tasks")
                for page in paginator.paginate(cluster=cluster_arn):
                    task_arns.extend(page.get("taskArns", []))
                if not task_arns:
                    continue
                for i in range(0, len(task_arns), 100):
                    batch = task_arns[i:i+100]
                    tasks = ecs.describe_tasks(cluster=cluster_arn, tasks=batch).get("tasks", [])
                    for task in tasks:
                        for attachment in task.get("attachments", []):
                            for detail in attachment.get("details", []):
                                if detail.get("name") == "networkInterfaceId":
                                    eni_id = detail.get("value")
                                    # ENI SG association already captured above via ENI scan
                                    break
        except Exception as e:
            logging.warning(f"[ECS] {region}: {e}")

        # ---- Build result rows ----
        for sg in sgs:
            sg_id = sg["GroupId"]
            sg_name = sg.get("GroupName", "")
            sg_desc = sg.get("Description", "")
            resources = list(sg_map.get(sg_id, set()))

            if not resources:
                # Orphaned SG
                orphaned_sgs.append([region, sg_id, sg_name, sg_desc])
                continue

            inbound_rules = collect_rules(sg, "inbound")
            outbound_rules = collect_rules(sg, "outbound")

            for res in resources:
                res_type, res_id, res_name, subnet_id, subnet_name, subnet_type = res
                for proto, port_str, cidr, ip_ver, desc in inbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc,
                                    res_type, res_id, res_name,
                                    subnet_name, subnet_id, subnet_type,
                                    proto, port_str, cidr, ip_ver, desc, "INBOUND"])
                for proto, port_str, cidr, ip_ver, desc in outbound_rules:
                    results.append([region, sg_id, sg_name, sg_desc,
                                    res_type, res_id, res_name,
                                    subnet_name, subnet_id, subnet_type,
                                    proto, port_str, cidr, ip_ver, desc, "OUTBOUND"])

    except EndpointConnectionError:
        logging.warning(f"Region {region} not available")
    except Exception as e:
        logging.error(f"[{region}] Unexpected error: {e}")

    return results, orphaned_sgs


# ------------------- Main -------------------

def main():
    # Use a single session only for bootstrapping (get account ID + regions)
    bootstrap_session = boto3.Session(profile_name=AWS_PROFILE)
    account_id = get_account_id(bootstrap_session)

    config = Config(retries={'max_attempts': 5, 'mode': 'adaptive'}, connect_timeout=5, read_timeout=10)
    ec2_client = bootstrap_session.client("ec2", config=config)

    INBOUND_FILE  = f"{account_id}_security_groups_inbound.xlsx"
    OUTBOUND_FILE = f"{account_id}_security_groups_outbound.xlsx"
    ORPHAN_FILE   = f"{account_id}_security_groups_orphaned.xlsx"

    regions = [r["RegionName"] for r in ec2_client.describe_regions()["Regions"]]
    logging.info(f"Scanning {len(regions)} regions for account {account_id}")

    inbound_data, outbound_data, orphaned_data = [], [], []

    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        future_to_region = {
            executor.submit(scan_region, region, AWS_PROFILE): region
            for region in regions
        }
        for future in as_completed(future_to_region):
            region = future_to_region[future]
            try:
                results, orphans = future.result()
                for row in results:
                    if row[-1] == "INBOUND":
                        inbound_data.append(row[:-1])
                    else:
                        outbound_data.append(row[:-1])
                orphaned_data.extend(orphans)
                logging.info(f"Completed region: {region} — {len(results)} rule rows, {len(orphans)} orphaned SGs")
            except Exception as e:
                logging.error(f"Error processing region {region}: {e}")

    columns = [
        "Region", "SG ID", "SG Name", "SG Description",
        "Resource Type", "Resource ID", "Resource Name",
        "Subnet Name", "Subnet ID", "Subnet Type",
        "Protocol", "Port(s)", "CIDR/Reference", "IP Version/Type", "Rule Description"
    ]
    orphan_columns = ["Region", "SG ID", "SG Name", "SG Description"]

    merge_cols = ["Region", "SG ID", "SG Name", "SG Description",
                  "Resource ID", "Resource Name", "Subnet Name", "Subnet ID"]

    for df, path in [
        (pd.DataFrame(inbound_data, columns=columns), INBOUND_FILE),
        (pd.DataFrame(outbound_data, columns=columns), OUTBOUND_FILE),
    ]:
        df.to_excel(path, index=False)
        apply_excel_formatting(path, merge_cols)
        logging.info(f"Saved: {path} ({len(df)} rows)")

    orphan_df = pd.DataFrame(orphaned_data, columns=orphan_columns)
    orphan_df.to_excel(ORPHAN_FILE, index=False)
    apply_excel_formatting(ORPHAN_FILE, ["Region"])
    logging.info(f"Saved: {ORPHAN_FILE} ({len(orphan_df)} orphaned SGs)")


if __name__ == "__main__":
    main()