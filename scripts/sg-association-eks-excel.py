import boto3
import pandas as pd
import concurrent.futures
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import sys

# CONFIG VARIABLES
AWS_PROFILE   = "aqcn"        # AWS CLI profile
AWS_REGION    = "us-east-1"            # e.g., "us-east-1"
ACCOUNT_ID    = "621924646669"            # AWS account ID
CLUSTER_NAME  = ""            # EKS cluster name
MAX_THREADS   = 10

INBOUND_FILE  = "<ACCOUNT_ID>_security_groups-(<CLUSTER_NAME>)-inbound.xlsx"
OUTBOUND_FILE = "<ACCOUNT_ID>_security_groups-(<CLUSTER_NAME>)-outbound.xlsx"

# HELPER FUNCTIONS
def input_with_default(prompt, default):
    val = input(f"{prompt} [{default}]: ").strip()
    return val if val else default

def create_session(profile, region):
    return boto3.Session(profile_name=profile, region_name=region)

def is_public_subnet(ec2, subnet_id, cache):
    try:
        if subnet_id in cache:
            return cache[subnet_id]
        rtbs = ec2.describe_route_tables(
            Filters=[{"Name":"association.subnet-id","Values":[subnet_id]}]
        ).get("RouteTables", [])
        if not rtbs:
            subnet = ec2.describe_subnets(SubnetIds=[subnet_id])["Subnets"][0]
            vpc_id = subnet["VpcId"]
            rtbs = ec2.describe_route_tables(Filters=[{"Name":"vpc-id","Values":[vpc_id]}])["RouteTables"]
        for rtb in rtbs:
            for route in rtb.get("Routes", []):
                if route.get("GatewayId","").startswith("igw-"):
                    cache[subnet_id] = "Public"
                    return "Public"
        cache[subnet_id] = "Private"
        return "Private"
    except Exception as e:
        print(f"[!] Error detecting subnet type for {subnet_id}: {e}")
        return "Unknown"

def merge_cells(workbook_path, sheet_name):
    try:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]
        merge_cols = [1,2,3,4,5,6,7]
        for col_idx in merge_cols:
            col_letter = get_column_letter(col_idx)
            start = 2
            prev = ws[f"{col_letter}2"].value
            for row in range(3, ws.max_row + 2):
                val = ws[f"{col_letter}{row}"].value
                if val != prev or row == ws.max_row + 1:
                    if row - 1 > start:
                        ws.merge_cells(f"{col_letter}{start}:{col_letter}{row-1}")
                        ws[f"{col_letter}{start}"].alignment = Alignment(vertical="top")
                    start = row
                    prev = val
        wb.save(workbook_path)
    except Exception as e:
        print(f"[!] Error merging cells in {workbook_path}: {e}")

def process_sg(ec2, sg_id, eni_info, subnet_cache):
    inbound, outbound = [], []
    try:
        sg = ec2.describe_security_groups(GroupIds=[sg_id])["SecurityGroups"][0]
        sg_name = sg["GroupName"]
        sg_desc = sg.get("Description","")
        subnet_id = eni_info["SubnetId"]
        subnet_type = is_public_subnet(ec2, subnet_id, subnet_cache)
        subnet_info = ec2.describe_subnets(SubnetIds=[subnet_id])["Subnets"][0]
        subnet_name = next((t["Value"] for t in subnet_info.get("Tags",[]) if t["Key"]=="Name"), "")
        resource_type = eni_info.get("InterfaceType","eni")
        resource_id = eni_info["NetworkInterfaceId"]
        resource_name = eni_info.get("Description","")

        def parse_rule(rule, direction):
            proto = rule.get("IpProtocol","all")
            port_from = rule.get("FromPort","")
            port_to = rule.get("ToPort","")
            port = f"{port_from}-{port_to}" if port_from != port_to else str(port_from)
            for ipr in rule.get("IpRanges",[]):
                row = [
                    AWS_REGION, sg_id, sg_name, sg_desc,
                    resource_type, resource_id, resource_name,
                    subnet_name, subnet_id, subnet_type,
                    proto, port, ipr.get("CidrIp",""), "IPv4", ipr.get("Description","")
                ]
                (inbound if direction=="in" else outbound).append(row)
            for ipr in rule.get("Ipv6Ranges",[]):
                row = [
                    AWS_REGION, sg_id, sg_name, sg_desc,
                    resource_type, resource_id, resource_name,
                    subnet_name, subnet_id, subnet_type,
                    proto, port, ipr.get("CidrIpv6",""), "IPv6", ipr.get("Description","")
                ]
                (inbound if direction=="in" else outbound).append(row)

        for r in sg.get("IpPermissions",[]):
            parse_rule(r,"in")
        for r in sg.get("IpPermissionsEgress",[]):
            parse_rule(r,"out")
    except Exception as e:
        print(f"[!] Error processing SG {sg_id}: {e}")
    return inbound, outbound

# MAIN SCRIPT
def main():
    global AWS_PROFILE, AWS_REGION, ACCOUNT_ID, CLUSTER_NAME, INBOUND_FILE, OUTBOUND_FILE

    print("\nGeneric EKS SG Scanner\n")

    AWS_PROFILE = input_with_default("AWS CLI profile", AWS_PROFILE)
    AWS_REGION  = input_with_default("AWS region", "us-east-1")
    ACCOUNT_ID  = input_with_default("AWS Account ID", ACCOUNT_ID)
    CLUSTER_NAME= input_with_default("EKS Cluster Name", CLUSTER_NAME)

    INBOUND_FILE  = INBOUND_FILE.replace("<ACCOUNT_ID>", ACCOUNT_ID).replace("<CLUSTER_NAME>", CLUSTER_NAME)
    OUTBOUND_FILE = OUTBOUND_FILE.replace("<ACCOUNT_ID>", ACCOUNT_ID).replace("<CLUSTER_NAME>", CLUSTER_NAME)

    print(f"Profile: {AWS_PROFILE}")
    print(f"Region: {AWS_REGION}")
    print(f"Account: {ACCOUNT_ID}")
    print(f"Cluster: {CLUSTER_NAME}")
    print(f"Inbound file:  {INBOUND_FILE}")
    print(f"Outbound file: {OUTBOUND_FILE}\n")

    session = create_session(AWS_PROFILE, AWS_REGION)
    ec2 = session.client("ec2")
    eks = session.client("eks")

    try:
        cluster_info = eks.describe_cluster(name=CLUSTER_NAME)["cluster"]
        sg_ids = set(cluster_info["resourcesVpcConfig"]["securityGroupIds"])
    except Exception as e:
        print(f"[!] Error fetching cluster {CLUSTER_NAME}: {e}")
        return

    try:
        for ng in eks.list_nodegroups(clusterName=CLUSTER_NAME)["nodegroups"]:
            desc = eks.describe_nodegroup(clusterName=CLUSTER_NAME,nodegroupName=ng)["nodegroup"]
            sg_ids.update(desc["resources"]["securityGroups"])
    except Exception as e:
        print(f"[!] Error fetching nodegroups: {e}")

    try:
        enis = ec2.describe_network_interfaces(Filters=[{"Name":"group-id","Values":list(sg_ids)}])["NetworkInterfaces"]
        print(f"Found {len(enis)} ENIs attached to EKS SGs\n")
    except Exception as e:
        print(f"[!] Error fetching ENIs: {e}")
        return

    inbound_all, outbound_all = [], []
    subnet_cache = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = []
        for eni in enis:
            for sg in eni["Groups"]:
                futures.append(executor.submit(process_sg, ec2, sg["GroupId"], eni, subnet_cache))
        for i, fut in enumerate(concurrent.futures.as_completed(futures), 1):
            try:
                inb, outb = fut.result()
                inbound_all.extend(inb)
                outbound_all.extend(outb)
            except Exception as e:
                print(f"[!] Error in future result: {e}")
            if i % 10 == 0 or i == len(futures):
                print(f"Processed {i}/{len(futures)} SGs...")

    columns = [
        "Region","SG ID","SG Name","SG Description","Resource Type","Resource ID","Resource Name",
        "Subnet Name","Subnet ID","Subnet Type","Protocol","Port(s)","CIDR/Reference","IP Version/Type","Rule Description"
    ]

    try:
        pd.DataFrame(inbound_all, columns=columns).to_excel(INBOUND_FILE, index=False)
        pd.DataFrame(outbound_all, columns=columns).to_excel(OUTBOUND_FILE, index=False)
        print("Excel files exported successfully.")
    except Exception as e:
        print(f"[!] Error exporting Excel files: {e}")

    merge_cells(INBOUND_FILE, "Sheet1")
    merge_cells(OUTBOUND_FILE, "Sheet1")
    print(f"Export complete")
    print(f"Inbound:  {INBOUND_FILE}")
    print(f"Outbound: {OUTBOUND_FILE}\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nAborted by user.")
        sys.exit(0)
